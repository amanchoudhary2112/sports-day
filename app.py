from flask import Flask, session, render_template, request, redirect, url_for, flash, send_file
import os
import openpyxl
from datetime import datetime

app = Flask(__name__)
app.secret_key = 'aman_key'

EXCEL_FILE = 'participants.xlsx'

# Ensure Excel file is initialized
if not os.path.exists(EXCEL_FILE):
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = "Participants"
    sheet.append("name", "email", "sport_selected", "mobile", "role", "course")
    wb.save(EXCEL_FILE)

# -------------------- ROUTES -------------------- #

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        if username == 'admin' and password == 'admin123':
            session['admin'] = True
            return redirect(url_for('participants'))
        else:
            flash('Invalid credentials', 'error')
    return render_template('login.html')


@app.route('/')
def index():
    sports = {
        'Badminton': '/static/images/badminton.webp',
        'Carrom Board': '/static/images/carrom board.jpg',
        'Chess': '/static/images/chess.webp',
        'Table Tennis': '/static/images/table tennis.jpg',
        'Tug Of War': '/static/images/tug of war.avif',
        'Lemon Spoon Race': '/static/images/lemon spoon race.jpg'
    }
    return render_template('index.html', sports=sports)


@app.route('/register', defaults={'sport': None}, methods=['GET', 'POST'])
@app.route('/register/<sport>', methods=['GET', 'POST'])
def register(sport):
    sports_list = ['Badminton', 'Carrom Board', 'Chess', 'Table Tennis', 'Tug Of War','Lemon Race']

    if request.method == 'POST':
        name = request.form['name']
        email = request.form['email']
        sport_selected = request.form['sport']
        mobile = request.form.get('mobile', '')
        role = request.form.get('role', '')
        course = request.form.get('course', '')

        wb = openpyxl.load_workbook(EXCEL_FILE)
        sheet = wb['Participants']
        sheet.append([name, email, sport_selected, mobile, role, course])

        wb.save(EXCEL_FILE)

        return redirect(url_for('success', name=name, sport=sport_selected))

    return render_template('register.html', sports=sports_list, selected_sport=sport)


@app.route('/success')
def success():
    name = request.args.get('name')
    sport = request.args.get('sport')
    return render_template('success.html', name=name, sport=sport)

@app.route('/gallery')
def gallery():
    # Example static list of image URLs (you can later use dynamic DB values)
    sports_images = [
        "https://source.unsplash.com/800x600/?sports-day,athlete",
        "https://source.unsplash.com/801x601/?relay-race",
        "https://source.unsplash.com/802x602/?school-sports,celebration",
        "https://source.unsplash.com/803x603/?tug-of-war",
        "https://source.unsplash.com/804x604/?running,track",
        "https://source.unsplash.com/805x605/?prize-distribution" ,
        "https://source.unsplash.com/806x606/?sports-team,celebration",
        "https://source.unsplash.com/807x607/?school-sports-day",
        "https://source.unsplash.com/808x608/?track-and-field",
    ]

    return render_template('gallery.html', images=sports_images)


@app.route('/participants')
def participants():
    if not session.get('admin'):
        return redirect(url_for('login'))

    search_query = request.args.get('search', '').lower()

    wb = openpyxl.load_workbook(EXCEL_FILE)
    sheet = wb['Participants']
    all_rows = list(sheet.iter_rows(values_only=True))[1:]

    if search_query:      
        participants = [p for p in all_rows if any(search_query in str(field).lower() for field in p)]
    else:
        participants = all_rows

    return render_template('participants.html', participants=participants, current_year=datetime.now().year, search_query=search_query)



@app.route('/download')
def download_excel():
    if not session.get('admin'):  # Fixed: use 'admin' to match login session
        return redirect(url_for('login'))
    return send_file(EXCEL_FILE, as_attachment=True)

# Store last deleted row in session
@app.route('/delete/<int:row_id>', methods=['POST'])
def delete_participant(row_id):
    if not session.get('admin'):
        return redirect(url_for('login'))

    wb = openpyxl.load_workbook(EXCEL_FILE)
    sheet = wb['Participants']
    deleted_row = list(sheet.iter_rows(values_only=True))[row_id]  # Exclude header
    session['last_deleted'] = deleted_row
    sheet.delete_rows(row_id + 2)  # +2 because row 0 = header, row 1 = index 1
    wb.save(EXCEL_FILE)
    flash("Participant deleted.", "info")
    return redirect(url_for('participants'))

@app.route('/delete_all', methods=['POST'])
def delete_all():
    if not session.get('admin'):
        return redirect(url_for('login'))

    wb = openpyxl.load_workbook(EXCEL_FILE)
    sheet = wb['Participants']
    all_rows = list(sheet.iter_rows(values_only=True))[1:]  # Exclude header
    session['last_deleted_all'] = all_rows
    sheet.delete_rows(2, sheet.max_row)
    wb.save(EXCEL_FILE)
    flash("All participants deleted.", "info")
    return redirect(url_for('participants'))

@app.route('/undo_delete')
def undo_delete():
    if not session.get('admin'):
        return redirect(url_for('login'))

    wb = openpyxl.load_workbook(EXCEL_FILE)
    sheet = wb['Participants']
    if 'last_deleted' in session:
        sheet.append(session['last_deleted'])
        session.pop('last_deleted')
        flash("Last deleted participant restored.", "success")
    elif 'last_deleted_all' in session:
        for row in session['last_deleted_all']:
            sheet.append(row)
        session.pop('last_deleted_all')
        flash("All deleted participants restored.", "success")
    else:
        flash("Nothing to undo.", "warning")
    wb.save(EXCEL_FILE)
    return redirect(url_for('participants'))
@app.route('/edit/<int:row_id>', methods=['GET', 'POST'])
def edit_participant(row_id):
    if not session.get('admin'):
        return redirect(url_for('login'))

    wb = openpyxl.load_workbook(EXCEL_FILE)
    sheet = wb['Participants']
    data = list(sheet.iter_rows(values_only=True))
    if request.method == 'POST':
        updated = [
            request.form['name'],
            request.form['email'],
            request.form['sport'],
            request.form['mobile'],
            request.form['role'],
            request.form['course']
        ]
        for col, val in enumerate(updated, start=1):
            sheet.cell(row=row_id+2, column=col, value=val)
        wb.save(EXCEL_FILE)
        flash('Participant updated successfully.', 'success')
        return redirect(url_for('participants'))

    participant = data[row_id + 1]
    return render_template('edit.html', participant=participant, row_id=row_id)




@app.route('/logout')
def logout():
    session.pop('admin', None)
    return redirect(url_for('index'))




# -------------------- MAIN -------------------- #

if __name__ == '__main__':
    app.run(debug=True)
